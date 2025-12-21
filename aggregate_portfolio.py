import pandas as pd
import os
import glob
from openpyxl import Workbook

# Configuration
INPUT_FOLDER = r"D:\Bhardwaj\Antigravity_Projects\MF_Data_Compiler\PPFAS_2025_Disclosures"
OUTPUT_FILE = r"D:\Bhardwaj\Antigravity_Projects\MF_Data_Compiler\PPFAS_Equity_Analysis.xlsx"

COLUMN_MAPPING = {
    "Name of the Instrument": "Name",
    "ISIN": "ISIN",
    "Industry/Rating": "Rating",
    "Quantity": "Quantity",
    "Market value(Rs. in Lakhs)": "MarketValue",
    "% to Net Assets": "PctAssets"
}

def normalize_header(header):
    if not isinstance(header, str): return ""
    h = header.lower().replace('\n', ' ').strip()
    if "instrument" in h: return "Name"
    if "isin" in h: return "ISIN"
    if "industry" in h or "rating" in h: return "Rating"
    if "quantity" in h: return "Quantity"
    if "market value" in h or "rs. in lakhs" in h: return "MarketValue"
    if "% to net assets" in h or "percentage" in h or "%" in h: return "PctAssets"
    return header

def read_portfolio_file(filepath):
    try:
        # Read roughly to find header
        df_raw = pd.read_excel(filepath, engine='xlrd', header=None)
        
        # 1. Find Column Header
        header_row_idx = -1
        for i, row in df_raw.iterrows():
            row_vals = [str(x).lower() for x in row.values]
            if any("isin" in v for v in row_vals) and any("quantity" in v for v in row_vals):
                header_row_idx = i
                break
        
        if header_row_idx == -1: return None
        
        # Reload with header. Note: Data starts after this.
        df = pd.read_excel(filepath, engine='xlrd', header=header_row_idx)
        
        # Rename columns
        new_cols = {}
        for col in df.columns:
            cleaned = normalize_header(col)
            if cleaned in COLUMN_MAPPING.values():
                new_cols[col] = cleaned
        df.rename(columns=new_cols, inplace=True)
        
        # 2. Filter Rows for specific section
        # We need to iterate and look for section markers.
        # The 'Name' column usually contains the markers. `df['Name']`.
        
        start_marker = "(a) Listed / awaiting listing on Stock Exchanges"
        stop_markers = ["Arbitrage", "Sub Total", "Total", "Debt", "Cash", "Grand Total"]
        
        filtered_rows = []
        capturing = False
        
        for idx, row in df.iterrows():
            # Check row content for markers
            # Get content from Name column or join all columns if Name is NaN/missing
            content_check = str(row.get('Name', ''))
            if not content_check or content_check.lower() == 'nan':
                 content_check = " ".join([str(x) for x in row.values])
                 
            # Aggressively clean: remove multiple spaces, newlines
            content_lower = " ".join(content_check.lower().split())
            
            # Use partial match for start marker
            # Note: July file has "Stock Exchange" (singular) vs "Stock Exchanges" (plural)
            if "(a) listed" in content_lower and "stock" in content_lower:
                capturing = True
                continue
            
            if capturing:
                if any(m.lower() in content_lower for m in stop_markers):
                    capturing = False
                    break 
                
                # Add valid data row if it has ISIN
                isin = row.get('ISIN')
                if pd.notna(isin) and str(isin).strip() != '' and str(isin).lower() != 'nan':
                    filtered_rows.append(row)
                    
        return pd.DataFrame(filtered_rows)

    except Exception as e:
        print(f"Error reading {os.path.basename(filepath)}: {e}")
        return None

def main():
    files = glob.glob(os.path.join(INPUT_FOLDER, "*.xls"))
    if not files:
        print("No files found.")
        return

    print("Aggregating Equity data...")
    
    portfolio = {}
    all_months = set()
    
    for f in files:
        fname = os.path.basename(f).replace(".xls", "").replace(".xlsx", "")
        all_months.add(fname)
        print(f"Processing {fname}...")
        
        df = read_portfolio_file(f)
        if df is None: continue
        
        for _, row in df.iterrows():
            isin = str(row['ISIN']).strip()
            
            if isin not in portfolio:
                portfolio[isin] = {
                    'Name': row.get('Name'),
                    'Rating': row.get('Rating'),
                    'Months': {}
                }
            
            # Update name if previously missing
            if not portfolio[isin]['Name'] and pd.notna(row.get('Name')):
                portfolio[isin]['Name'] = row.get('Name')
                
            portfolio[isin]['Months'][fname] = {
                'Quantity': row.get('Quantity'),
                'MarketValue': row.get('MarketValue'),
                'PctAssets': row.get('PctAssets')
            }

    # Sort months
    month_order = ['January', 'February', 'March', 'April', 'May', 'June', 
                   'July', 'August', 'September', 'October', 'November', 'December']
    
    def sort_months_key(m_str):
        try:
            if '_' in m_str:
                m, y = m_str.split('_')
                return (int(y) if y.isdigit() else 9999, month_order.index(m) if m in month_order else 99)
            return (9999, 99)
        except: return (9999, 99)

    sorted_months = sorted(list(all_months), key=sort_months_key)
    
    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Equity Analysis"
    
    # Import Styles
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    # 1. Main Title (Row 1)
    # We will merge this later once we know max_col
    ws.cell(1, 1, "PPFAS Flexi cap Fund")
    
    # 2. Static info (Row 2)
    headers_static = ["Name of the Instrument", "ISIN", "Industry/Rating"]
    for i, h in enumerate(headers_static, 1):
        ws.cell(2, i, h)
    
    # Headers (Row 2 & 3 now)
    col_idx = 4
    column_totals = {} # key: col_idx, value: sum
    
    for m in sorted_months:
        # Tier 1 Header (Month) at Row 2
        c = ws.cell(2, col_idx, m)
        ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx+2)
        
        # Tier 2 Headers at Row 3
        ws.cell(3, col_idx, "Quantity")
        ws.cell(3, col_idx+1, "Market Value (Rs. Lakhs)")
        ws.cell(3, col_idx+2, "% Net Assets")
        
        # Init totals
        column_totals[col_idx] = 0.0   # Qty
        column_totals[col_idx+1] = 0.0 # Val
        column_totals[col_idx+2] = 0.0 # Pct
        
        col_idx += 3
        
    final_col_idx = col_idx - 1
    
    # Merge Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=final_col_idx)

    # Data Rows (Start at Row 4)
    row_idx = 4
    for isin, data in portfolio.items():
        ws.cell(row_idx, 1, data['Name'])
        ws.cell(row_idx, 2, isin)
        ws.cell(row_idx, 3, data['Rating'])
        
        col_idx = 4
        for m in sorted_months:
            m_data = data['Months'].get(m, {})
            
            # Helper safely parses and returns value to write
            def get_val(key):
                try: 
                    v = float(m_data.get(key)) 
                    return v
                except: return 0.0
            
            qty = get_val('Quantity')
            val = get_val('MarketValue')
            pct = get_val('PctAssets')
            
            # Write to cell (write None if 0 to check if that looks better? No, 0 is fine, or blank?)
            # Usually strict 0 might be clutter. Let's write value if it exists in data, else None.
            # But we need to sum zeroes.
            
            # Logic: If month data missing, it's effectively 0 for sum. 
            # Write the raw value from dict if present, else empty.
            
            node_qty = m_data.get('Quantity')
            node_val = m_data.get('MarketValue')
            node_pct = m_data.get('PctAssets')
            
            def write_val(c_idx, raw_val, sum_key):
                v_float = 0.0
                try: v_float = float(raw_val)
                except: pass
                
                # Add to total
                column_totals[c_idx] += v_float
                
                if raw_val is not None:
                    ws.cell(row_idx, c_idx, v_float)
                else:
                    ws.cell(row_idx, c_idx, None)

            write_val(col_idx, node_qty, 'Quantity')
            write_val(col_idx+1, node_val, 'MarketValue')
            write_val(col_idx+2, node_pct, 'PctAssets')
            
            col_idx += 3
        row_idx += 1

    # Totals Row
    total_row = row_idx
    ws.cell(total_row, 1, "Total")
    
    # Write sums
    for c_idx, total_val in column_totals.items():
        ws.cell(total_row, c_idx, total_val)

    # Apply Formatting
    
    # Styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(wrap_text=True, vertical='center', horizontal='center')
    middle_align = Alignment(vertical='center')

    # Border Styles
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                         
    # Custom border generator for month separators
    # We want a medium border on the RIGHT of the % column (every 3rd dynamic col)
    # Col indices: 1, 2, 3 (Static) | 4, 5, 6 (M1) | 7, 8, 9 (M2) ...
    # So if (col_idx - 3) % 3 == 0 -> medium right
    
    def get_border(c_idx):
        # Base is thin all around
        left = Side(style='thin')
        right = Side(style='thin')
        top = Side(style='thin')
        bottom = Side(style='thin')
        
        # Check for Month Block Separator
        # Month blocks end at 6, 9, 12... 
        # (c_idx) % 3 == 0?  No, static is 3 cols.
        # Month 1: 4,5,6. End at 6. (6-3)%3 == 0. Correct.
        # Static: 1,2,3. End at 3. (3-3)%3 == 0. Correct.
        
        if c_idx >= 3 and (c_idx - 3) % 3 == 0:
            right = Side(style='medium')
            
        return Border(left=left, right=right, top=top, bottom=bottom)

    # Apply Title Style (Row 1)
    for cell in ws[1]:
        cell.font = title_font
        cell.alignment = center_align

    # Apply Header Styles (Row 2 & 3)
    # Row 2 is Months (Center)
    for cell in ws[2]:
        cell.font = header_font
        cell.alignment = center_align
        cell.border = get_border(cell.column)
        
    # Row 3 is Sub-headers (Wrap)
    for cell in ws[3]:
        cell.font = header_font
        # Force wrap alignment for subheaders
        cell.alignment = wrap_align
        cell.border = get_border(cell.column)

    # Apply Data Styles (Row 4 to Total Row)
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Formats
    fmt_indian_int = "[>=10000000]##\,##\,##\,##0;[>=100000]##\,##\,##0;##,##0"
    fmt_indian_float = "[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00"
    fmt_pct = "0.00%"
    
    for row in ws.iter_rows(min_row=4, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.alignment = middle_align
            cell.border = get_border(cell.column)
            
            # Bold the Total row
            if cell.row == total_row:
                cell.font = Font(bold=True)
            
            # Number Formats
            c_idx = cell.column
            if c_idx >= 4:
                rem = (c_idx - 4) % 3
                if rem == 0: # Quantity
                    cell.number_format = fmt_indian_int
                elif rem == 1: # Value
                    cell.number_format = fmt_indian_float
                elif rem == 2: # Pct
                    cell.number_format = fmt_pct
                    
    # Column Widths
    from openpyxl.utils import get_column_letter
    
    # Auto-fit simple logic
    for col_i in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_i)
        max_length = 0
        
        # Iterate rows in this column
        for row_i in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_i, column=col_i)
            try:
                # Skip if cell is part of merge but not top-left?
                # Actually value is safely accessible usually, might be None
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
                
        adjusted_width = (max_length + 2)
        if adjusted_width > 50: adjusted_width = 50
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(OUTPUT_FILE)
    print(f"Saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
